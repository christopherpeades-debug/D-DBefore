"""Build armor_enchants.json and shield_enchants.json from spreadsheet + d20 SRD."""
from __future__ import annotations

import ast
import json
import re
import urllib.request
from html import unescape
from pathlib import Path

OUT_DIR = Path(__file__).resolve().parent
SRD_URL = "https://www.d20srd.org/srd/magicItems/magicArmor.htm"
XLSX_CANDIDATES = (
    OUT_DIR / "armor_shield_enchants_temp.xlsx",
    Path(r"C:\Users\refle\OneDrive\Desktop\Armor and sheild Enchants.xlsx"),
    Path(r"C:\Users\refle\Downloads\armor_shield_enchants.xlsx"),
)
SPREADSHEET_FALLBACK = OUT_DIR / "spreadsheet_enchants_rows.json"
SHEET_CANDIDATES = ("Armor & Shield Enchants", "Sheet1")

# h5 ids for specific magic armors/shields — not special-ability enchants.
SRD_SPECIFIC_ITEM_IDS = {
    "adamantineBreastplate", "bandedMailofLuck", "breastplateofCommand", "celestialArmor",
    "demonArmor", "dragonhidePlate", "dwarvenPlate", "elvenChain", "mithralFullPlateofSpeed",
    "mithralShirt", "plateArmoroftheDeep", "rhinoHide", "absorbingShield", "castersShield",
    "darkwoodBuckler", "darkwoodShield", "lionsShield", "mithralHeavyShield", "spinedShield",
    "wingedShield",
}

SKILL_NAMES = [
    "Appraise", "Balance", "Bluff", "Climb", "Concentration", "Craft", "Decipher Script",
    "Diplomacy", "Disable Device", "Disguise", "Escape Artist", "Forgery",
    "Gather Information", "Handle Animal", "Heal", "Hide", "Intimidate", "Jump",
    "Listen", "Move Silently", "Open Lock", "Perform", "Profession", "Ride", "Search",
    "Sense Motive", "Sleight of Hand", "Spellcraft", "Spot", "Survival", "Swim",
    "Tumble", "Use Magic Device", "Use Rope",
    "Knowledge (arcana)", "Knowledge (architecture and engineering)",
    "Knowledge (dungeoneering)", "Knowledge (geography)", "Knowledge (history)",
    "Knowledge (local)", "Knowledge (nature)", "Knowledge (nobility and royalty)",
    "Knowledge (religion)", "Knowledge (the planes)",
]

SPELL_PHRASES = [
    ("greater teleport", "Teleport, Greater"),
    ("dimension door", "Dimension Door"),
    ("greater invisibility", "Greater Invisibility"),
    ("death ward", "Death Ward"),
    ("displacement", "Displacement"),
    ("ethereal jaunt", "Ethereal Jaunt"),
    ("etherealness", "Etherealness"),
    ("freedom of movement", "Freedom of Movement"),
    ("water breathing", "Water Breathing"),
    ("stoneskin", "Stoneskin"),
    ("daylight", "Daylight"),
    ("blur", "Blur"),
    ("blink", "Blink"),
    ("fly", "Fly"),
    ("haste", "Haste"),
    ("invisibility", "Invisibility"),
    ("polymorph", "Polymorph"),
    ("teleport", "Teleport"),
    ("darkvision", "Darkvision"),
    ("feather fall", "Feather Fall"),
    ("gust of wind", "Gust of Wind"),
    ("fog cloud", "Fog Cloud"),
    ("spider climb", "Spider Climb"),
    ("water walk", "Water Walk"),
    ("meld into stone", "Meld into Stone"),
    ("resist energy", "Resist Energy"),
    ("silence", "Silence"),
    ("grease", "Grease"),
    ("limited wish", "Limited Wish"),
    ("miracle", "Miracle"),
]

RESISTANCE_AMOUNTS = {
    "": 10,
    ", Improved": 20,
    ", Greater": 30,
}

ENERGY_TYPES = ("Acid", "Cold", "Electricity", "Fire", "Sonic")


def slug(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(text or "").lower()).strip()


def title_name(raw: str) -> str:
    s = str(raw or "").strip()
    if not s:
        return s
    if "," in s:
        head, tail = s.split(",", 1)
        return f"{head.strip().title()}, {tail.strip().title()}"
    return " ".join(part.capitalize() if part.isupper() else part.title() for part in s.split())


def parse_cost(cost_raw: str) -> dict:
    s = str(cost_raw or "").strip().replace(",", "")
    m = re.search(r"\+\s*(\d+(?:\.\d+)?)\s*bonus", s, re.I)
    if m:
        return {"price_bonus": int(float(m.group(1)))}
    m = re.search(r"\+\s*(\d+(?:\.\d+)?)\s*gp", s, re.I)
    if m:
        return {"flat_gp": int(float(m.group(1)))}
    return {}


def classify_applies(applies_raw: str) -> tuple[str, list[str]]:
    s = str(applies_raw or "").strip().lower()
    restrictions: list[str] = []
    if "armor or shield" in s:
        base = "both"
    elif "metal armor" in s:
        base = "armor"
        restrictions.append("metal_armor")
    elif "shield" in s and "armor" not in s:
        base = "shield"
    elif "armor" in s and "shield" not in s:
        base = "armor"
    elif "shield" in s:
        base = "both"
    else:
        base = "armor"
    if "light armor" in s:
        restrictions.append("light_armor")
    if "tower" in s:
        restrictions.append("no_tower_shield")
    return base, restrictions


def clean_html_text(html_fragment: str) -> str:
    text = re.sub(r"<br\s*/?>", "\n", html_fragment, flags=re.I)
    text = re.sub(r"</p>", "\n", text, flags=re.I)
    text = re.sub(r"<[^>]+>", " ", text)
    text = unescape(text).replace("\xa0", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n\s*\n+", "\n\n", text)
    return text.strip()


def resolve_skill(token: str) -> str | None:
    folded = str(token or "").strip().lower()
    for skill in SKILL_NAMES:
        if skill.lower() == folded:
            return skill
    return None


def parse_effects(description: str, *, extra_text: str = "") -> dict:
    blob = f"{description}\n{extra_text}"
    effects: dict = {}
    skill_bonus: dict[str, int] = {}
    for match in re.finditer(
        r"([+-]?\d+)\s+"
        r"(?:(competence|resistance|morale|enhancement|circumstance|insight)\s+)?"
        r"(?:bonus|penalty)\s+on\s+"
        r"([A-Za-z ()]+?)\s+checks?",
        blob,
        re.I,
    ):
        skill = resolve_skill(match.group(3).strip())
        if skill:
            skill_bonus[skill] = skill_bonus.get(skill, 0) + int(match.group(1))
    if skill_bonus:
        effects["skill_bonus"] = skill_bonus

    save_map = {"reflex": "reflex_save", "fortitude": "fort_save", "will": "will_save"}
    for match in re.finditer(
        r"([+-]?\d+)\s+(?:(resistance|enhancement|morale|competence|insight)\s+)?"
        r"(?:bonus\s+)?on\s+(Reflex|Fortitude|Will)\s+saving\s+throws?",
        blob,
        re.I,
    ):
        key = save_map[match.group(3).lower()]
        effects[key] = max(int(effects.get(key, 0) or 0), int(match.group(1)))

    defenses = []
    seen = set()
    for match in re.finditer(r"resistance\s+(\d+)\s+to\s+(\w+)", blob, re.I):
        entry = ("Resistance", match.group(1), match.group(2).title())
        if entry not in seen:
            seen.add(entry)
            defenses.append({"type": "Resistance", "value": match.group(1), "desc": match.group(2).title()})
    for match in re.finditer(r"(fire|cold|acid|electricity|sonic)\s+resistance\s+(\d+)", blob, re.I):
        entry = ("Resistance", match.group(2), match.group(1).title())
        if entry not in seen:
            seen.add(entry)
            defenses.append({"type": "Resistance", "value": match.group(2), "desc": match.group(1).title()})
    for match in re.finditer(r"absorbs the first\s+(\d+)\s+points of\s+(\w+)\s+damage", blob, re.I):
        entry = ("Resistance", match.group(1), match.group(2).title())
        if entry not in seen:
            seen.add(entry)
            defenses.append({"type": "Resistance", "value": match.group(1), "desc": match.group(2).title()})
    for match in re.finditer(r"(?:gain\s+)?DR\s+(\d+)\s*/\s*([^.;,\n]+)", blob, re.I):
        val = f"{match.group(1)}/{match.group(2).strip()}"
        entry = ("DR", val, "")
        if entry not in seen:
            seen.add(entry)
            defenses.append({"type": "DR", "value": val, "desc": ""})
    for match in re.finditer(r"spell resistance\s+(?:is\s+)?(\d+)", blob, re.I):
        entry = ("SR", match.group(1), "")
        if entry not in seen:
            seen.add(entry)
            defenses.append({"type": "SR", "value": match.group(1), "desc": ""})
    for label, pct in (("light fortification", 25), ("moderate fortification", 75), ("heavy fortification", 100)):
        if label in blob.lower():
            entry = ("Fortification", f"{pct}%", label.title())
            if entry not in seen:
                seen.add(entry)
                defenses.append({"type": "Fortification", "value": f"{pct}%", "desc": label.title()})
    if "25% chance" in blob.lower() and "fortification" in blob.lower():
        entry = ("Fortification", "25%", "Light Fortification")
        if entry not in seen:
            seen.add(entry)
            defenses.append({"type": "Fortification", "value": "25%", "desc": "Light Fortification"})
    if defenses:
        effects["defenses"] = defenses
    return effects


def spell_phrase_in_text(needle: str, text: str) -> bool:
    needle = str(needle or "").strip().lower()
    if not needle:
        return False
    return re.search(rf"\b{re.escape(needle)}\b", str(text or ""), re.I) is not None


def parse_abilities(description: str) -> dict:
    abilities: dict = {"uses_per_day": 0, "max_charges": 0, "granted_spells": []}
    if re.search(r"any\s+number\s+of\s+times\s+per\s+day", description, re.I):
        abilities["uses_per_day"] = 0
    elif re.search(r"\bonce\s+per\s+day\b", description, re.I):
        abilities["uses_per_day"] = 1
    else:
        match = re.search(r"(\d+)\s*/\s*day", description, re.I)
        if match:
            abilities["uses_per_day"] = int(match.group(1))

    grants: list[str] = []
    for pattern in (
        r"\(\s*as\s+(?:the\s+)?(.+?)\s+spell\s*\)",
        r"\(\s*as\s+(?:the\s+)?(.+?)\s*\)",
    ):
        for match in re.finditer(pattern, description, re.I):
            phrase = match.group(1).strip().rstrip(".")
            for needle, spell in SPELL_PHRASES:
                if spell_phrase_in_text(needle, phrase) and spell not in grants:
                    grants.append(spell)
    # Ignore prerequisite / comparison mentions (e.g. "similar to the resist energy spell").
    scrubbed = re.sub(
        r"similar to the\s+.+?\s+spell",
        "",
        description,
        flags=re.I,
    )
    scrubbed = re.sub(
        r"prerequisite[s]?:.+$",
        "",
        scrubbed,
        flags=re.I,
    )
    for needle, spell in sorted(SPELL_PHRASES, key=lambda pair: -len(pair[0])):
        if spell_phrase_in_text(needle, scrubbed) and spell not in grants:
            grants.append(spell)
    abilities["granted_spells"] = grants
    return abilities


def merge_effects_dict(base: dict, extra: dict) -> dict:
    merged = dict(base)
    for key, value in (extra or {}).items():
        if key == "defenses":
            existing = merged.setdefault("defenses", [])
            for defense in value:
                if defense not in existing:
                    existing.append(defense)
        elif key == "skill_bonus":
            skill_map = merged.setdefault("skill_bonus", {})
            for skill, amount in value.items():
                skill_map[skill] = max(skill_map.get(skill, 0), amount)
        else:
            merged[key] = max(int(merged.get(key, 0) or 0), int(value or 0))
    return merged


def finalize_entry(entry: dict) -> dict:
    description = str(entry.get("description", "") or "").strip()
    notes = str(entry.get("notes", "") or "").strip()
    effects = parse_effects(description, extra_text=notes)
    if entry.get("effects"):
        effects = merge_effects_dict(effects, entry["effects"])
    if effects:
        entry["effects"] = effects
    else:
        entry.pop("effects", None)

    abilities = parse_abilities(description)
    if entry.get("abilities"):
        stored = entry["abilities"]
        if stored.get("granted_spells"):
            abilities["granted_spells"] = list(stored["granted_spells"])
        for key in ("uses_per_day", "max_charges"):
            if stored.get(key) is not None:
                abilities[key] = stored[key]
    has_abilities = (
        abilities.get("granted_spells")
        or int(abilities.get("uses_per_day") or 0) > 0
        or int(abilities.get("max_charges") or 0) > 0
    )
    if has_abilities:
        entry["abilities"] = abilities
    else:
        entry.pop("abilities", None)
    return entry


def fetch_srd_html() -> str:
    req = urllib.request.Request(SRD_URL, headers={"User-Agent": "DnDBesideEnchantBuilder/1.0"})
    with urllib.request.urlopen(req, timeout=60) as resp:
        return resp.read().decode("utf-8", errors="replace")


def parse_srd_table_applies(html: str) -> tuple[set[str], set[str]]:
    armor_names: set[str] = set()
    shield_names: set[str] = set()
    for table_id, target in (("tableArmorSpecialAbilities", armor_names), ("tableShieldSpecialAbilities", shield_names)):
        match = re.search(rf'<table[^>]*id="{table_id}"[^>]*>(.*?)</table>', html, re.I | re.S)
        if not match:
            continue
        for row in re.findall(r"<tr[^>]*>(.*?)</tr>", match.group(1), re.I | re.S)[1:]:
            cells = re.findall(r"<t[dh][^>]*>(.*?)</t[dh]>", row, re.I | re.S)
            if len(cells) < 2:
                continue
            ability = clean_html_text(cells[-2])
            if not ability or ability.lower().startswith("roll twice"):
                continue
            target.add(title_name(ability))
    return armor_names, shield_names


def srd_applies_to(name: str, armor_names: set[str], shield_names: set[str]) -> str:
    in_armor = name in armor_names
    in_shield = name in shield_names
    if in_armor and in_shield:
        return "both"
    if in_shield:
        return "shield"
    return "armor"


def parse_srd_notes(body: str) -> str:
    parts = []
    for pattern, label in (
        (r"(faint|moderate|strong)\s+[^;]+;\s*CL\s*([^;]+);", "caster"),
        (r"Craft Magic Arms and Armor[^;]*", "prereq"),
        (r"Price[^.]*\.", "price"),
    ):
        match = re.search(pattern, body, re.I)
        if not match:
            continue
        if label == "caster":
            parts.append(f"Caster Level: {match.group(2).strip()}")
            parts.append(f"{match.group(1).title()} aura.")
        elif label == "prereq":
            parts.append(f"Prerequisites: {clean_html_text(match.group(0))}")
        elif label == "price":
            parts.append(clean_html_text(match.group(0)))
    return " ".join(parts)


def parse_srd_cost(body: str) -> dict:
    result = {}
    match = re.search(r"Price\s*\+?\s*(\d+(?:,\d{3})*)\s*gp", body, re.I)
    if match:
        result["flat_gp"] = int(match.group(1).replace(",", ""))
    match = re.search(r"Price\s*\+(\d+)\s*bonus", body, re.I)
    if match:
        result["price_bonus"] = int(match.group(1))
    match = re.search(r"SR\s*(\d+)\).*?\+(\d+)\s*bonus", body, re.I)
    if match:
        result["price_bonus"] = int(match.group(2))
        result.setdefault("effects", {}).setdefault("defenses", []).append(
            {"type": "SR", "value": match.group(1), "desc": ""},
        )
    return result


def heading_to_name(heading: str) -> str:
    heading = clean_html_text(heading)
    if "," in heading:
        head, tail = heading.split(",", 1)
        return f"{head.strip()}, {tail.strip()}"
    return heading.strip()


def scrape_srd_enchants(html: str) -> dict[str, dict]:
    armor_names, shield_names = parse_srd_table_applies(html)
    entries: dict[str, dict] = {}
    start = html.find("Magic Armor and Shield Special Ability Descriptions")
    chunk = html[start:] if start >= 0 else html
    sections = re.findall(r'<h5 id="([^"]+)">(.*?)</h5>(.*?)(?=<h5 id=)', chunk, re.I | re.S)
    descriptions_by_name: dict[str, str] = {}

    for section_id, heading_html, body_html in sections:
        if section_id in SRD_SPECIFIC_ITEM_IDS:
            continue
        name = heading_to_name(heading_html)
        if section_id == "ghostTouchArmor":
            name = "Ghost Touch"
        body = clean_html_text(body_html)
        paragraphs = [p.strip() for p in body.split("\n\n") if p.strip()]
        description = paragraphs[0] if paragraphs else body
        descriptions_by_name[name] = description
        entry = {
            "description": description,
            "applies_to": srd_applies_to(name, armor_names, shield_names),
        }
        notes = parse_srd_notes(body)
        if notes:
            entry["notes"] = notes
        entry.update(parse_srd_cost(body))
        entries[name] = entry

    # Expand multi-tier SRD entries not fully covered by single headings.
    extras: list[tuple[str, dict]] = []

    for variant, pct, bonus in (
        ("Fortification, Light", "25%", 1),
        ("Fortification, Moderate", "75%", 3),
        ("Fortification, Heavy", "100%", 5),
    ):
        extras.append((
            variant,
            {
                "description": (
                    f"This suit of armor or shield negates critical hits and sneak attacks "
                    f"{pct} of the time (as fortification)."
                ),
                "applies_to": "both",
                "price_bonus": bonus,
                "effects": {"defenses": [{"type": "Fortification", "value": pct, "desc": variant}]},
                "notes": entries.get("Fortification", {}).get("notes", ""),
            },
        ))

    for sr_value, bonus in ((13, 2), (15, 3), (17, 4), (19, 5)):
        extras.append((
            f"Spell Resistance ({sr_value})",
            {
                "description": (
                    f"This armor or shield grants spell resistance {sr_value} while worn."
                ),
                "applies_to": "both",
                "price_bonus": bonus,
                "effects": {"defenses": [{"type": "SR", "value": str(sr_value), "desc": ""}]},
                "notes": entries.get("Spell Resistance", {}).get("notes", ""),
            },
        ))

    for energy in ENERGY_TYPES:
        for suffix, amount, flat_gp in (("", 10, 18000), (", Improved", 20, 42000), (", Greater", 30, 66000)):
            base_name = f"{energy} Resistance{suffix}"
            if base_name in entries:
                continue
            base_desc = descriptions_by_name.get(f"{energy} Resistance", "")
            if suffix:
                description = f"As {energy.lower()} resistance, except it absorbs the first {amount} points of {energy.lower()} damage per attack."
            else:
                description = base_desc or (
                    f"The armor or shield absorbs the first {amount} points of {energy.lower()} damage per attack."
                )
            extras.append((
                base_name,
                {
                    "description": description,
                    "applies_to": srd_applies_to(base_name, armor_names, shield_names),
                    "flat_gp": flat_gp,
                    "effects": {
                        "defenses": [{"type": "Resistance", "value": str(amount), "desc": energy}],
                    },
                    "notes": entries.get(f"{energy} Resistance{suffix}", {}).get("notes", "")
                    or entries.get(f"{energy} Resistance", {}).get("notes", ""),
                },
            ))

    for base, improved_bonus, greater_bonus in (
        ("Shadow", 10, 15),
        ("Silent Moves", 10, 15),
        ("Slick", 10, 15),
    ):
        skill_map = {
            "Shadow": "Hide",
            "Silent Moves": "Move Silently",
            "Slick": "Escape Artist",
        }
        skill = skill_map[base]
        if base in entries:
            entries[base].setdefault("effects", {}).setdefault("skill_bonus", {})[skill] = 5
        for label, bonus, flat_gp in (("Improved", improved_bonus, 15000), ("Greater", greater_bonus, 33750)):
            name = f"{base}, {label}"
            if name in entries:
                entries[name].setdefault("effects", {}).setdefault("skill_bonus", {})[skill] = bonus
                continue
            extras.append((
                name,
                {
                    "description": f"As {base.lower()}, except it grants a +{bonus} competence bonus on {skill} checks.",
                    "applies_to": "armor",
                    "flat_gp": flat_gp,
                    "effects": {"skill_bonus": {skill: bonus}},
                    "notes": entries.get(base, {}).get("notes", ""),
                },
            ))

    for name, payload in extras:
        if name not in entries:
            entries[name] = payload

    # Remove generic spell resistance / fortification headings in favor of expanded variants.
    entries.pop("Spell Resistance", None)
    entries.pop("Fortification", None)

    finalized = {}
    for name, entry in entries.items():
        finalized[name] = finalize_entry(dict(entry))
    return finalized


def load_spreadsheet_rows() -> list[tuple]:
    for candidate in XLSX_CANDIDATES:
        if not candidate.exists():
            continue
        try:
            from openpyxl import load_workbook
            wb = load_workbook(candidate, data_only=True)
        except PermissionError:
            print(f"Skipping locked file: {candidate}")
            continue
        ws = None
        for sheet_name in SHEET_CANDIDATES:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                break
        if ws is None:
            ws = wb.active
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 4 or not row or not row[0]:
                continue
            rows.append(tuple(row[:5]))
        if rows:
            print(f"Loaded {len(rows)} spreadsheet rows from {candidate.name}")
            return rows

    if SPREADSHEET_FALLBACK.exists():
        rows = [tuple(row) for row in json.loads(SPREADSHEET_FALLBACK.read_text(encoding="utf-8"))]
        print(f"Loaded {len(rows)} spreadsheet rows from {SPREADSHEET_FALLBACK.name}")
        return rows

    session_rows = []
    compaction = Path(
        r"C:\Users\refle\.grok\sessions\C%3A%5CUsers%5Crefle"
        r"\019ee0e3-a339-76e0-983a-72b92d989689\compaction_requests"
        r"\4dca1700-5bd6-4218-93bc-7bb7484a663d.json"
    )
    if compaction.exists():
        text = compaction.read_text(encoding="utf-8", errors="replace")
        seen = set()
        for match in re.finditer(
            r"\((?:'(?:[^'\\]|\\.)*'|\"(?:[^\"\\]|\\.)*\")(?:,\s*(?:'(?:[^'\\]|\\.)*'|\"(?:[^\"\\]|\\.)*\"|None))*\)",
            text,
        ):
            try:
                row = ast.literal_eval(match.group(0))
            except (SyntaxError, ValueError):
                continue
            if not isinstance(row, tuple) or len(row) < 5 or not row[0]:
                continue
            if row[0] in ("Name", "D&D 3.5 Magical Armor and Shield Enchantments"):
                continue
            if "Compiled from" in str(row[0]):
                continue
            if row[0] in seen:
                continue
            seen.add(row[0])
            session_rows.append(row[:5])
    if session_rows:
        SPREADSHEET_FALLBACK.write_text(
            json.dumps([list(row) for row in session_rows], indent=2, ensure_ascii=False) + "\n",
            encoding="utf-8",
        )
        print(f"Recovered {len(session_rows)} spreadsheet rows from session logs")
        return session_rows
    return []


def build_spreadsheet_entries(rows: list[tuple]) -> dict[str, dict]:
    entries: dict[str, dict] = {}
    for raw_name, applies_raw, cost_raw, desc, notes in rows:
        name = title_name(raw_name)
        applies_to, restrictions = classify_applies(applies_raw)
        entry = {
            "description": str(desc or "").strip(),
            "applies_to": applies_to,
        }
        if restrictions:
            entry["restrictions"] = restrictions
        if notes:
            entry["notes"] = str(notes).strip()
        entry.update(parse_cost(cost_raw))
        entries[name] = finalize_entry(entry)
    return entries


def merge_entries(*sources: dict[str, dict]) -> dict[str, dict]:
    merged: dict[str, dict] = {}
    slug_map: dict[str, str] = {}

    def attach(name: str, entry: dict, *, prefer_existing_description: bool = False):
        key = slug(name)
        if key in slug_map and slug_map[key] != name:
            # Keep the more specific name already stored.
            existing_name = slug_map[key]
            current = merged[existing_name]
            for field, value in entry.items():
                if field not in current or not current[field]:
                    current[field] = value
            return
        slug_map[key] = name
        if name not in merged:
            merged[name] = dict(entry)
            return
        current = merged[name]
        if prefer_existing_description:
            old_desc = str(current.get("description", ""))
            new_desc = str(entry.get("description", ""))
            if len(new_desc) > len(old_desc):
                current["description"] = new_desc
        for field, value in entry.items():
            if field in ("effects", "abilities"):
                continue
            if field not in current or not current[field]:
                current[field] = value
        if entry.get("effects"):
            current_effects = current.setdefault("effects", {})
            for eff_key, eff_val in entry["effects"].items():
                if eff_key == "defenses":
                    existing = current_effects.setdefault("defenses", [])
                    for defense in eff_val:
                        if defense not in existing:
                            existing.append(defense)
                elif eff_key == "skill_bonus":
                    skill_map = current_effects.setdefault("skill_bonus", {})
                    for skill, amount in eff_val.items():
                        skill_map[skill] = max(skill_map.get(skill, 0), amount)
                else:
                    current_effects[eff_key] = max(int(current_effects.get(eff_key, 0) or 0), int(eff_val or 0))
        if entry.get("abilities"):
            current["abilities"] = entry["abilities"]

    for source in sources:
        for name, entry in source.items():
            attach(name, entry)

    finalized = {}
    for name, entry in merged.items():
        finalized[name] = finalize_entry(entry)
    return dict(sorted(finalized.items(), key=lambda kv: kv[0].lower()))


def split_by_applies(merged: dict[str, dict]) -> tuple[dict, dict]:
    armor_db: dict[str, dict] = {}
    shield_db: dict[str, dict] = {}
    for name, entry in merged.items():
        applies = entry.get("applies_to", "both")
        payload = dict(entry)
        if applies in ("armor", "both"):
            armor_db[name] = payload
        if applies in ("shield", "both"):
            shield_db[name] = payload
    return armor_db, shield_db


def main():
    print("Fetching d20 SRD magic armor special abilities...")
    html = fetch_srd_html()
    srd_entries = scrape_srd_enchants(html)
    print(f"Parsed {len(srd_entries)} SRD special abilities")

    sheet_rows = load_spreadsheet_rows()
    sheet_entries = build_spreadsheet_entries(sheet_rows) if sheet_rows else {}
    if sheet_entries:
        print(f"Parsed {len(sheet_entries)} spreadsheet enchants")

    merged = merge_entries(sheet_entries, srd_entries)
    armor_db, shield_db = split_by_applies(merged)

    for path, db in (
        (OUT_DIR / "armor_enchants.json", armor_db),
        (OUT_DIR / "shield_enchants.json", shield_db),
    ):
        path.write_text(json.dumps(db, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
        print(f"Wrote {len(db)} entries to {path.name}")


if __name__ == "__main__":
    main()